"""
AMBROSIO — Interface Web (Streamlit) — versão autossuficiente
Execute com: streamlit run app.py
"""

import re
import tempfile
import pandas as pd
import streamlit as st
from pathlib import Path
from datetime import datetime
from rapidfuzz import process, fuzz
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Funções de leitura ─────────────────────────────────────────────────────────

def limpar_valor(texto):
    if not texto:
        return None
    texto = re.sub(r"[^\d,]", "", str(texto))
    texto = texto.replace(",", ".")
    partes = texto.split(".")
    if len(partes) > 2:
        texto = "".join(partes[:-1]) + "." + partes[-1]
    try:
        return float(texto)
    except ValueError:
        return None

def extrair_pdf(caminho_pdf, coluna_valor):
    registros = []
    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina in pdf.pages:
            tabelas = pagina.extract_tables()
            for tabela in tabelas:
                for linha in tabela:
                    if not linha:
                        continue
                    celulas = [c for c in linha if c and str(c).strip()]
                    if len(celulas) < 2:
                        continue
                    favorecido = str(celulas[0]).strip()
                    for celula in reversed(celulas):
                        valor = limpar_valor(str(celula))
                        if valor and valor > 0:
                            registros.append({"favorecido": favorecido, coluna_valor: valor})
                            break
    df = pd.DataFrame(registros) if registros else pd.DataFrame(columns=["favorecido", coluna_valor])
    return df[df[coluna_valor] > 0].reset_index(drop=True)

# ── Funções de comparação ──────────────────────────────────────────────────────

TOLERANCIA = 0.01
LIMIAR = 80

def normalizar(nome):
    return " ".join(str(nome).upper().split())

def comparar(df_lote, df_itau):
    resultados = []
    candidatos = df_itau["favorecido"].tolist()
    usados = set()

    for _, linha in df_lote.iterrows():
        fav = linha["favorecido"]
        val_lote = linha["valor_lote"]
        resultado = process.extractOne(normalizar(fav), [normalizar(c) for c in candidatos], scorer=fuzz.token_sort_ratio)

        match = None
        score = 0
        if resultado and resultado[1] >= LIMIAR:
            score = resultado[1]
            idx_list = df_itau[df_itau["favorecido"].apply(normalizar) == normalizar(candidatos[resultado[2]])].index.tolist()
            idx = next((i for i in idx_list if i not in usados), None)
            if idx is not None:
                usados.add(idx)
                val_itau = df_itau.loc[idx, "valor_itau"]
                match = df_itau.loc[idx, "favorecido"]
                dif = round(val_lote - val_itau, 2)
                status = "OK" if abs(dif) <= TOLERANCIA else "DIVERGENTE"
                resultados.append({"favorecido_lote": fav, "valor_lote": val_lote, "favorecido_itau": match, "valor_itau": val_itau, "diferenca": dif, "status": status, "score": score})
                continue

        resultados.append({"favorecido_lote": fav, "valor_lote": val_lote, "favorecido_itau": None, "valor_itau": None, "diferenca": val_lote, "status": "NÃO ENCONTRADO NO ITAÚ", "score": 0})

    for idx, linha in df_itau.iterrows():
        if idx not in usados:
            resultados.append({"favorecido_lote": None, "valor_lote": None, "favorecido_itau": linha["favorecido"], "valor_itau": linha["valor_itau"], "diferenca": -linha["valor_itau"], "status": "NÃO ENCONTRADO NO LOTE", "score": 0})

    return pd.DataFrame(resultados)

# ── Funções de relatório ───────────────────────────────────────────────────────

def fmt_moeda(v):
    if v is None or pd.isna(v):
        return "-"
    return f"R$ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")

def gerar_excel(df):
    COR_CAB = "1F3864"; COR_OK = "C6EFCE"; COR_DIV = "FFCCCC"; COR_NA = "FFEB9C"
    BORDA = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    def cor(status):
        return {"OK": COR_OK, "DIVERGENTE": COR_DIV}.get(status, COR_NA)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"

    total = len(df); ok = len(df[df.status=="OK"]); div = len(df[df.status=="DIVERGENTE"]); na = total-ok-div
    dif_total = df["diferenca"].fillna(0).sum()

    ws["A1"] = "AMBROSIO — Relatório de Divergências"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", start_color=COR_CAB)
    ws.merge_cells("A1:C1")
    ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    for i, (l, v) in enumerate([("Total",total),("✅ OK",ok),("⚠️ Divergentes",div),("❓ Não encontrados",na),("Diferença total",fmt_moeda(dif_total))], 4):
        ws[f"A{i}"] = l; ws[f"B{i}"] = v; ws[f"A{i}"].font = Font(bold=True)

    ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 20

    cols = ["Favorecido (Lote)","Valor Lote","Favorecido (Itaú)","Valor Itaú","Diferença","Status"]
    wd = wb.create_sheet("Detalhamento")

    for ci, t in enumerate(cols, 1):
        c = wd.cell(1, ci, t); c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", start_color=COR_CAB); c.alignment = Alignment(horizontal="center"); c.border = BORDA

    for ri, (_, ln) in enumerate(df.iterrows(), 2):
        fill = PatternFill("solid", start_color=cor(ln.status))
        for ci, v in enumerate([ln.get("favorecido_lote") or "-", fmt_moeda(ln.get("valor_lote")), ln.get("favorecido_itau") or "-", fmt_moeda(ln.get("valor_itau")), fmt_moeda(ln.get("diferenca")), ln.get("status","")], 1):
            c = wd.cell(ri, ci, v); c.border = BORDA; c.fill = fill

    for ci, w in enumerate([35,18,35,18,18,28], 1):
        wd.column_dimensions[get_column_letter(ci)].width = w

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name

# ── Interface ──────────────────────────────────────────────────────────────────

st.set_page_config(page_title="AMBROSIO", page_icon="💼", layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.header-box { background: linear-gradient(135deg, #1F3864, #2E5490); border-radius: 12px; padding: 2rem 2.5rem; margin-bottom: 2rem; color: white; }
.header-box h1 { font-size: 2.2rem; font-weight: 700; margin: 0; }
.header-box p { opacity: 0.8; margin: 0.4rem 0 0 0; }
.metric-card { background: white; border-radius: 10px; padding: 1.2rem 1.5rem; box-shadow: 0 1px 4px rgba(0,0,0,0.08); border-left: 4px solid #1F3864; margin-bottom: 1rem; }
.metric-card.ok { border-left-color: #22c55e; } .metric-card.error { border-left-color: #ef4444; } .metric-card.warn { border-left-color: #f59e0b; }
.metric-label { font-size: 0.78rem; color: #6b7280; font-weight: 600; text-transform: uppercase; }
.metric-value { font-size: 1.8rem; font-weight: 700; color: #111827; }
.stButton>button { background: linear-gradient(135deg, #1F3864, #2E5490); color: white; border: none; border-radius: 8px; padding: 0.6rem 2rem; font-weight: 600; width: 100%; }
.chat-user { background: #eff6ff; border-radius: 12px 12px 4px 12px; padding: 0.8rem 1.2rem; margin: 0.5rem 0; text-align: right; color: #1e40af; font-weight: 500; }
.chat-bot { background: #f1f5f9; border-radius: 12px 12px 12px 4px; padding: 0.8rem 1.2rem; margin: 0.5rem 0; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="header-box"><h1>💼 AMBROSIO</h1><p>Conferência automática de pagamentos · DB1 Group</p></div>', unsafe_allow_html=True)

if "df_resultado" not in st.session_state:
    st.session_state.df_resultado = None
if "chat" not in st.session_state:
    st.session_state.chat = []

st.subheader("📤 Carregar arquivos")
c1, c2 = st.columns(2)
with c1:
    st.markdown("**Lote de Pagamentos**")
    arq_lote = st.file_uploader("Lote", type=["pdf"], key="lote", label_visibility="collapsed")
    if arq_lote: st.success(f"✔ {arq_lote.name}")
with c2:
    st.markdown("**Extrato Itaú**")
    arq_itau = st.file_uploader("Itau", type=["pdf"], key="itau", label_visibility="collapsed")
    if arq_itau: st.success(f"✔ {arq_itau.name}")

st.markdown("")
_, cb, _ = st.columns([2,1,2])
with cb:
    processar = st.button("⚖️ Processar", disabled=not (arq_lote and arq_itau))

if processar:
    with st.spinner("Lendo e comparando..."):
        with tempfile.TemporaryDirectory() as tmp:
            p_lote = Path(tmp) / arq_lote.name
            p_itau = Path(tmp) / arq_itau.name
            p_lote.write_bytes(arq_lote.read())
            p_itau.write_bytes(arq_itau.read())
            df_lote = extrair_pdf(p_lote, "valor_lote")
            df_itau = extrair_pdf(p_itau, "valor_itau")
            st.session_state.df_resultado = comparar(df_lote, df_itau)
            st.session_state.chat = []

if st.session_state.df_resultado is not None:
    df = st.session_state.df_resultado
    total = len(df); ok = len(df[df.status=="OK"]); div = len(df[df.status=="DIVERGENTE"]); na = total-ok-div
    dif = df["diferenca"].fillna(0).sum()

    st.markdown("---")
    st.subheader("📊 Resumo")
    m1,m2,m3,m4 = st.columns(4)
    with m1: st.markdown(f'<div class="metric-card ok"><div class="metric-label">Sem divergência</div><div class="metric-value">{ok}</div></div>', unsafe_allow_html=True)
    with m2: st.markdown(f'<div class="metric-card error"><div class="metric-label">Divergentes</div><div class="metric-value">{div}</div></div>', unsafe_allow_html=True)
    with m3: st.markdown(f'<div class="metric-card warn"><div class="metric-label">Não encontrados</div><div class="metric-value">{na}</div></div>', unsafe_allow_html=True)
    with m4: st.markdown(f'<div class="metric-card {"error" if abs(dif)>0.01 else "ok"}"><div class="metric-label">Diferença total</div><div class="metric-value">{fmt_moeda(dif)}</div></div>', unsafe_allow_html=True)

    def fmt_df(df_in):
        d = df_in.copy()
        for col in ["valor_lote","valor_itau","diferenca"]:
            if col in d.columns: d[col] = d[col].apply(lambda v: fmt_moeda(v))
        return d.rename(columns={"favorecido_lote":"Favorecido (Lote)","valor_lote":"Valor Lote","favorecido_itau":"Favorecido (Itaú)","valor_itau":"Valor Itaú","diferenca":"Diferença","status":"Status","score":"Similaridade %"}).fillna("-")

    st.markdown("---")
    t1, t2 = st.tabs(["📋 Todos os registros","⚠️ Divergências"])
    with t1: st.dataframe(fmt_df(df), use_container_width=True, height=400)
    with t2:
        df_div = df[df.status != "OK"]
        if df_div.empty: st.success("✅ Nenhuma divergência encontrada!")
        else: st.dataframe(fmt_df(df_div), use_container_width=True, height=400)

    st.markdown("---")
    st.subheader("⬇️ Baixar relatório")
    caminho_excel = gerar_excel(df)
    with open(caminho_excel, "rb") as f:
        st.download_button("📥 Baixar relatorio_divergencias.xlsx", f.read(), "relatorio_divergencias.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.markdown("---")
    st.subheader("💬 Pergunte sobre os pagamentos")
    st.caption("Ex: 'Qual fornecedor teve maior diferença?' · 'Quantos ficaram pendentes?'")

    for msg in st.session_state.chat:
        if msg["role"] == "user": st.markdown(f'<div class="chat-user">🧑 {msg["content"]}</div>', unsafe_allow_html=True)
        else: st.markdown(f'<div class="chat-bot">🤖 {msg["content"]}</div>', unsafe_allow_html=True)

    with st.form("chat_form", clear_on_submit=True):
        pergunta = st.text_input("Pergunta", placeholder="Ex: qual foi a maior divergência?", label_visibility="collapsed")
        enviar = st.form_submit_button("Enviar")

    if enviar and pergunta.strip():
        st.session_state.chat.append({"role": "user", "content": pergunta})
        import anthropic
        ctx = df.to_string(index=False)
        resumo = f"Total: {total}. OK: {ok}. Divergentes: {div}. Não encontrados: {na}. Diferença total: {fmt_moeda(dif)}."
        cliente = anthropic.Anthropic()
        resp = cliente.messages.create(
            model="claude-sonnet-4-6", max_tokens=1000,
            system=f"Você é o AMBROSIO, assistente financeiro da DB1 Group. Responda em português sobre os dados abaixo.\nResumo: {resumo}\nDados:\n{ctx}",
            messages=[{"role":"user","content":pergunta}]
        )
        st.session_state.chat.append({"role":"assistant","content":resp.content[0].text})
        st.rerun()
else:
    st.info("📂 Carregue os dois PDFs acima e clique em **Processar** para começar.")
